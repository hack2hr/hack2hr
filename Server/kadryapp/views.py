from django.shortcuts import render
import statsmodels
import json
from django.http.response import JsonResponse
from django.http.response import HttpResponse
from bson import json_util
from django.views.decorators.csrf import csrf_exempt
from pymongo import MongoClient
from bson.objectid import ObjectId
import pprint
from datetime import datetime
import openpyxl
from openpyxl import Workbook
client = MongoClient('localhost', 27017)
db = client['KadryTest']


# query = Coll.objects.filter(ids="1").values()
    #query = Coll.objects.filter(id="1").values()
    #print(query)
    # usercX = request.GET.get("var") ?var=qewwqe
    # title = request.POST.get("title")

def index(request):
    response = JsonResponse(json_util.dumps({"test":123 }), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response
 
@csrf_exempt
def testPost(request):
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    body_unicode = request.body.decode('utf-8')
    jsonValue = json.loads(body_unicode)
    #content = body['content']
    print(jsonValue)
    response = JsonResponse(json_util.dumps({"test": jsonValue["title"]}), safe = False)
    #jsonValue["title"] get json value
    response["Access-Control-Allow-Origin"] = "*"
    return response

@csrf_exempt 
def testGet(request):
    testVal = request.GET.get("testVal")
    response = JsonResponse(json_util.dumps({"test": testVal}), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response

@csrf_exempt 
def testAddMongo(request):
    data = {
        "year": "2000",
        "totalyear": "1000",
        "data": {
            "q1": {
                "totalq1": "800",
                "workAble": "200",
                "migrants": "200",
                "other": {
                    "old": "200",
                    "young": "200"
                }
            },
            "q2": {
                "totalq2": "800",
                "workAble": "200",
                "migrants": "200",
                "other": {
                    "old": "200",
                    "young": "200"
                }
            },
            "q3": {
                "totalq3": "800",
                "workAble": "200",
                "migrants": "200",
                "other": {
                    "old": "200",
                    "young": "200"
                }
            },
            "q4": {
                "totalq4": "800",
                "workAble": "200",
                "migrants": "200",
                "other": {
                    "old": "200",
                    "young": "200"
                }
            }        
        }
    }
    
    collection = db['People']
    doc_id = collection.insert_one(data).inserted_id
    document = collection.find_one({'_id': ObjectId(doc_id)})
    response = JsonResponse(json_util.dumps(document), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response

@csrf_exempt 
def testPyMongo(request):
    from Server.core.models import Model
   
    collection = db['People']
    test = []
    for doc in collection.find():
        test.append(doc)
    

    response["Access-Control-Allow-Origin"] = "*"    
    response = JsonResponse(json_util.dumps(test), safe = False)
    return response



#People

#Get all People documents
@csrf_exempt 
#http://10.0.0.4:8080/api/people/all/
def apiPeopleAll(request):
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    collection = db['People']
    test = []
    for doc in collection.find():
        test.append(doc)
    response = JsonResponse(json_util.dumps(test), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response

#Get one People document
@csrf_exempt 
#http://10.0.0.4:8080/api/people/one/?year=int
def apiPeopleOne(request):
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    year = request.GET.get("year")
    collection = db['People']  
    result = collection.find_one({"year": year})  
    response = JsonResponse(json_util.dumps(result), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response

#Insert People document
@csrf_exempt 
def apiPeopleUpdate(request):
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    body_unicode = request.body.decode('utf-8')
    jsonValue = json.loads(body_unicode)

    

    year = jsonValue['year']
    totalyear = jsonValue['totalyear']
    madeby = jsonValue['madeby']
    auto = jsonValue['auto']
    accepted = jsonValue['accepted']

    q1totalq1 = jsonValue['data']['q1']['totalq1']
    q1workAble = jsonValue['data']['q1']['workAble']
    q1migrants = jsonValue['data']['q1']['migrants']
    q1old = jsonValue['data']['q1']['other']['old']
    q1young = jsonValue['data']['q1']['other']['young']

    q2totalq2 = jsonValue['data']['q2']['totalq4']
    q2workAble = jsonValue['data']['q2']['workAble']
    q2migrants = jsonValue['data']['q2']['migrants']
    q2old = jsonValue['data']['q2']['other']['young']
    q2young = jsonValue['data']['q2']['other']['young']

    q3totalq3 = jsonValue['data']['q3']['totalq3']
    q3workAble = jsonValue['data']['q3']['workAble']
    q3migrants = jsonValue['data']['q3']['migrants']
    q3old = jsonValue['data']['q3']['other']['young']
    q3young = jsonValue['data']['q3']['other']['young']

    q4totalq4 = jsonValue['data']['q4']['totalq4']
    q4workAble = jsonValue['data']['q4']['workAble']
    q4migrants = jsonValue['data']['q4']['migrants']
    q4old = jsonValue['data']['q4']['other']['young']
    q4young = jsonValue['data']['q4']['other']['young']

    data = { 
        "$set": {
            "year": year,
            "totalyear": totalyear,
            "madeby": madeby,
            "auto": auto,
            "accepted": accepted,
            "data": {
                "q1": {
                    "totalq1": q1totalq1,
                    "workAble": q1workAble,
                    "migrants": q1migrants,
                    "other": {
                        "old": q1old,
                        "young": q1young
                    }
                },
                "q2": {
                    "totalq2": q2totalq2,
                    "workAble": q2workAble,
                    "migrants": q2migrants,
                    "other": {
                        "old": q2old,
                        "young": q2young
                    }
                },
                "q3": {
                    "totalq3": q3totalq3,
                    "workAble": q3workAble,
                    "migrants": q3migrants,
                    "other": {
                        "old": q3old,
                        "young": q3young
                    }
                },
                "q4": {
                    "totalq4": q4totalq4,
                    "workAble": q4workAble,
                    "migrants": q4migrants,
                    "other": {
                        "old": q4old,
                        "young": q4young
                    }
                }        
            }
        }
    }

    
    collection = db['People']
    myquery = { "year": year }

    collection.update_one(myquery, data)
    
    response = JsonResponse(json_util.dumps(document), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response



#Staff

#Get all Staff documents
@csrf_exempt 
#http://10.0.0.4:8080/api/staff/all/
def apiStaffAll(request):
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    collection = db['Staff']
    test = []
    for doc in collection.find():
        test.append(doc)
    response = JsonResponse(json_util.dumps(test), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response

#Get one Staff document
@csrf_exempt 
#http://10.0.0.4:8080/api/people/one/?_id=str
def apiStaffOne(request):
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    _id = request.GET.get("_id")
    collection = db['Staff']  
    result = collection.find_one({"_id": _id})  
    response = JsonResponse(json_util.dumps(result), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response

#Insert Staff document
@csrf_exempt 
#http://10.0.0.4:8080/api/staff/add/
def apiStaffAdd(request):
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    body_unicode = request.body.decode('utf-8')
    jsonValue = json.loads(body_unicode)
    surname = jsonValue['surname']
    name = jsonValue['name']
    secname = jsonValue['secname']
    email = jsonValue['email']
    rank = jsonValue['rank']
    isAdmin = jsonValue['isAdmin']

    data = {
        "surname": surname,
        "name": name,
        "secname": secname,
        "email": email,
        "rank": rank,
        "isAdmin": isAdmin
    }

    collection = db['Staff']
    doc_id = collection.insert_one(data).inserted_id
    response = JsonResponse(json_util.dumps(doc_id), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response

#Update Staff document
@csrf_exempt 
#http://10.0.0.4:8080/api/staff/update/
def apiStaffUpdate(request):
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    body_unicode = request.body.decode('utf-8')
    jsonValue = json.loads(body_unicode)
    _id = jsonValue['_id']
    surname = jsonValue['surname']
    name = jsonValue['name']
    secname = jsonValue['secname']
    email = jsonValue['email']
    rank = jsonValue['rank']
    isAdmin = jsonValue['isAdmin']

    
    myquery = { "_id": ObjectId(""+_id+"") }
    data = { "$set": {
            "surname": surname,
            "name": name,
            "secname": secname,
            "email": email,
            "rank": rank,
            "isAdmin": isAdmin
        }
    }

    collection = db['Staff']
    collection.update_one(myquery, data)
    response = JsonResponse(json_util.dumps(myquery), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response

#Delete Staff document
@csrf_exempt 
#http://10.0.0.4:8080/api/staff/delete/
def apiStaffDelete(request):
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    body_unicode = request.body.decode('utf-8')
    jsonValue = json.loads(body_unicode)
    _id = jsonValue['_id']
    
    myquery = { "_id": ObjectId(""+_id+"") }
    
    collection = db['Staff']
    collection.delete_one(myquery)
    response = JsonResponse(json_util.dumps(myquery), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response


#Predict People
@csrf_exempt 
#http://10.0.0.4:8080/api/people/predict/
def apiPeoplePredict(request):
    from Server.core.models import Model
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    body_unicode = request.body.decode('utf-8')
    jsonValue = json.loads(body_unicode)
    year = jsonValue['year']
    yearsrange = jsonValue['yearsRange']
    dataY = jsonValue['dataY']
    dataX = jsonValue['dataX']
    modelValue = jsonValue['modelValue']
    

    collection = db['People']
    stringForModel = []
    for doc in collection.find():
        stringForModel.append(doc)
    
    data = {
        "current_year": year,
        "years_to_predict": yearsrange,
        "x_params": dataX,
        "y_param": dataY,
    }

    result = Model(stringForModel, data).predict(modelValue)



    
    response = JsonResponse(json_util.dumps(result), safe = False)
    response["Access-Control-Allow-Origin"] = "*"
    return response


#Predict People
@csrf_exempt 
#http://10.0.0.4:8080/api/people/predict/
def apiPeopleDownload(request):
    
    if(request.method == "OPTIONS"): 
        response = JsonResponse({})
        response["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response["Access-Control-Allow-Origin"] = "http://13.79.21.196:8080"
        response["Access-Control-Allow-Headers"] = "origin, x-requested-with, content-type, x-ijt"
        response["Access-Control-Allow-Origin"] = "*"
        return response
    collection = db['People']
    data = []
    for doc in collection.find():
        data.append(doc)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}---test.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d')
    )


    workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Data'

    columns = [
        'Год',
        'Численность трудовых ресурсов, в том числе:',
        'К1 Численность трудовых ресурсов, в том числе: ',
        'К1 трудоспособное население в трудоспособном возрасте',
        'К1 иностранные трудовые мигранты',
        'К1 лица старше трудоспособного возраста и подростки, занятые в экономике, в том числе',
        'К1 лица старше трудоспособного возраста',
        'К1 подростки',
        'К2 Численность трудовых ресурсов, в том числе: ',
        'К2 трудоспособное население в трудоспособном возрасте',
        'К2 иностранные трудовые мигранты',
        'К2 лица старше трудоспособного возраста и подростки, занятые в экономике, в том числе',
        'К2 лица старше трудоспособного возраста',
        'К2 подростки',
        'К3 Численность трудовых ресурсов, в том числе: ',
        'К3 трудоспособное население в трудоспособном возрасте',
        'К3 иностранные трудовые мигранты',
        'К3 лица старше трудоспособного возраста и подростки, занятые в экономике, в том числе',
        'К3 лица старше трудоспособного возраста',
        'К3 подростки',
        'К4 Численность трудовых ресурсов, в том числе: ',
        'К4 трудоспособное население в трудоспособном возрасте',
        'К4 иностранные трудовые мигранты',
        'К4 лица старше трудоспособного возраста и подростки, занятые в экономике, в том числе',
        'К4 лица старше трудоспособного возраста',
        'К4 подростки',
        'Отчет принят',
        'Расчитан автоматически',
        'Расчитан сотрудником',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all 
    for part in data:
        row_num += 1
        print()
        # Define the data for each cell in the row 
        row = [
            part['year'],
            part['totalyear'],
            part['data']['q1']['totalq1'],
            part['data']['q1']['workAble'],
            part['data']['q1']['migrants'],
            part['data']['q1']['other']['old'] + part['data']['q1']['other']['young'],
            part['data']['q1']['other']['old'],
            part['data']['q1']['other']['young'],
            part['data']['q2']['totalq2'],
            part['data']['q2']['workAble'],
            part['data']['q2']['migrants'],
            part['data']['q2']['other']['old'] + part['data']['q2']['other']['young'],
            part['data']['q2']['other']['old'],
            part['data']['q2']['other']['young'],
            part['data']['q3']['totalq3'],
            part['data']['q3']['workAble'],
            part['data']['q3']['migrants'],
            part['data']['q3']['other']['old'] + part['data']['q3']['other']['young'],
            part['data']['q3']['other']['old'],
            part['data']['q3']['other']['young'],
            part['data']['q4']['totalq4'],
            part['data']['q4']['workAble'],
            part['data']['q4']['migrants'],
            part['data']['q4']['other']['old'] + part['data']['q4']['other']['young'],
            part['data']['q4']['other']['old'],
            part['data']['q4']['other']['young'],
            part['accepted'],
            part['auto'],
            part['madeby'],

        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response

    